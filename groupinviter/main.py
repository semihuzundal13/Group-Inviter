import smtplib
import openpyxl

file = ("dep_formu29ekim.xlsx")
wb = openpyxl.load_workbook(file)
sheet = wb.active

subject = "İTÜ Girişimcilik Kulübü IT Departmanı WP davet linki"
massage = "GROUP INVITE LINK \n Bu mail python smtp modulü kullanılarak atılmştır ;) \n Semih Uzundal IT Direktörü"
content = "Subject: {0}\n\n{1}".format(subject, massage)

mail = smtplib.SMTP("smtp.live.com", 587)
mail.ehlo()
mail.starttls()
ymail = "semihuzundal_1907@hotmail.com"
ypassword = "ENTER YOUR PASSWORD HERE"
mail.login(ymail, ypassword)

a = 471
while a < 537:
    dep = sheet.cell(row=a, column=5).value
    dep_list = dep.split(",")
    for i in dep_list:
        for m in dep_list:
            if m == " IT Departmanı":
                amail = sheet.cell(row=a, column=4).value
                mail.sendmail("semihuzundal_1907@hotmail.com", amail, content.encode("utf-8"))
            elif m == "IT Departmanı":
                amail = sheet.cell(row=a, column=4).value
                mail.sendmail("semihuzundal_1907@hotmail.com", amail, content.encode("utf-8"))

    a = a+1
