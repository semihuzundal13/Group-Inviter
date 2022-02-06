import smtplib
import openpyxl

file = ("dep_formu.xlsx")
wb = openpyxl.load_workbook(file)
sheet = wb.active

a = 2
while a < 266:
    dep = sheet.cell(row=a, column=5).value
    dep_list = dep.split(",")
    for i in dep_list:
        if i == " IT Departmanı":
            print(sheet.cell(row=a, column=5).value)
        elif i == "IT Departmanı":
            print(sheet.cell(row=a, column=5).value)
    a += 1

    amail = sheet.cell(row=a, column=4).value
    mail.sendmail("suzundal7@gmail.com", amail, content.encode("utf-8"))